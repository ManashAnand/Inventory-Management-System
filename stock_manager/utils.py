from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import numbers
import logging
from functools import reduce
from io import BytesIO
from openpyxl import Workbook, load_workbook
from rest_framework.response import Response
from rest_framework import status
from django.http import FileResponse
from django.db import transaction
from datetime import datetime
import pytz

from .models import Item, ShopItem, User, Admin


def sanitize_price(value, *, default="0.00") -> Decimal:
    """
    Convert an arbitrary value (possibly '£12.30', '12,345.6', float, numpy number)
    into a Decimal quantized to 2 dp. Raises ValueError if it cannot be parsed.
    """
    if value is None or (isinstance(value, str) and value.strip() == ""):
        value = default
    if isinstance(value, str):
        s = value.strip()
        s = s.replace("£", "").replace(",", "").replace("\u00A0", "").strip()
        candidate = s
    elif isinstance(value, numbers.Number):
        candidate = str(value)
    else:
        candidate = str(value).strip()
    try:
        d = Decimal(candidate)
    except (InvalidOperation, ValueError) as e:
        raise ValueError(f"Invalid retail price value: {value!r}") from e

    # Reject NaN or infinite values which can silently pass through Decimal()
    try:
        if d.is_nan() or d.is_infinite():
            raise ValueError(f"Invalid retail price value (NaN/Infinite): {value!r}")
    except Exception:
        # Some Decimal subclasses or unusual inputs may not have these methods; re-raise as generic ValueError
        raise ValueError(f"Invalid retail price value: {value!r}")

    d = d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return d

logger = logging.getLogger(__name__)

try:
    from .custom_funcs import spreadsheet_convert

    HAS_SPREADSHEET_CONVERT = True
except ImportError:
    spreadsheet_convert = None
    HAS_SPREADSHEET_CONVERT = False


class SpreadsheetTools:

    def __init__(self, request=None):
        self.request = request
        self.user = request.user

    def get_related_field(self, obj, field_name):
        """
        Follow a chain of related fields (e.g. "shop_user__username") and return the resulting value.
        """
        try:
            return reduce(
                lambda o, attr: getattr(o, attr, None) if o else None,
                field_name.split("__"),
                obj,
            )
        except AttributeError:
            return ""

    def convert_custom_incoming_format(self, workbook):
        logger.info(
            "convert_custom_incoming_format called. HAS_SPREADSHEET_CONVERT=%s",
            HAS_SPREADSHEET_CONVERT,
        )
        # returns a Workbook or raises Exception
        if HAS_SPREADSHEET_CONVERT:
            try:
                logger.info("Calling convert_excel...")
                return spreadsheet_convert.convert_excel(workbook)
            except Exception as e:
                logger.error(f"Error in convert_excel: {e}", exc_info=True)
                raise Exception(f"Custom spreadsheet conversion failed: {e}")
        else:
            raise Exception(
                "A 'custom_funcs/spreadsheet_convert.py' file does not exist or could not be imported. This is fine unless you are trying to map a custom upload spreadsheet schema into the database."
            )

    def create_excel_workbook(self):
        """
        Generate an Excel workbook containing two sheets.
        - 'Warehouse Stock' for Items.
        - 'Shop Stock' for ShopItems.
        The user must be in either the 'managers' or 'shop_users' group.
        """
        workbook = Workbook()
        # Create the 'Warehouse Stock' sheet
        item_sheet = workbook.active
        item_sheet.title = "Warehouse Stock"
        item_fields = ["sku", "description", "retail_price", "quantity"]
        item_header = ["SKU", "Description", "Retail Price", "Quantity"]
        item_sheet.append(item_header)
        for item in Item.objects.filter(is_active=True).only(*item_fields):
            row_data = [getattr(item, field, "") for field in item_fields]
            item_sheet.append(row_data)

        # Create the 'Shop Stock' sheet
        shop_item_sheet = workbook.create_sheet(title="Shop Stock")
        shop_item_relation_fields = ["shop_user", "item"]
        shop_item_retrieved_fields = [
            "shop_user__username",
            "item__sku",
            "item__description",
            "item__retail_price",
            "quantity",
        ]
        shop_item_header = [
            "Shop User",
            "SKU",
            "Description",
            "Retail Price",
            "Quantity",
        ]
        shop_item_sheet.append(shop_item_header)

        # If user is not a manager, limit the queryset
        is_manager = self.user.groups.filter(name="managers").exists()
        queryset = ShopItem.objects.select_related(*shop_item_relation_fields).only(
            *shop_item_retrieved_fields
        )
        if not is_manager:
            queryset = queryset.filter(shop_user__username=self.user.username)

        for shop_item in queryset:
            row_data = [
                self.get_related_field(shop_item, field)
                for field in shop_item_retrieved_fields
            ]
            shop_item_sheet.append(row_data)
        return workbook

    def generate_excel_response(self):
        """
        Convert an Excel workbook into a Django FileResponse that triggers a download.
        """
        formatted_datetime = datetime.now(pytz.timezone("EUROPE/LONDON")).strftime(
            "%d%b%Y_%H%M%S%Z"
        )
        filename = f"SSM_DATA_{formatted_datetime}.xlsx"
        workbook = self.create_excel_workbook()
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return FileResponse(
            output,
            as_attachment=True,
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def field_changed(self, instance, field_name, new_value):
        """
        Check whether a field on an instance would change if updated with new_value.
        Uses the model field's to_python() to convert the new value.
        """
        try:
            field_obj = instance._meta.get_field(field_name)
            norm_new_value = field_obj.to_python(new_value)
        except Exception:
            norm_new_value = (
                new_value  # fallback if field not found or conversion fails
            )

        # Handle None vs. empty string
        old_value = getattr(instance, field_name)
        if old_value is None and norm_new_value in [None, ""]:
            return False
        return old_value != norm_new_value

    def cleanup_orphaned_shopitems(self):
        """
        Remove ShopItem rows where item is NULL or item_id points to a non-existent Item (sku).
        """
        # Delete ShopItem rows where item is NULL
        ShopItem.objects.filter(item__isnull=True).delete()
        # Delete ShopItem rows where item_id points to a non-existent Item.sku (using ORM)
        orphaned_shopitems = ShopItem.objects.exclude(item__isnull=True).exclude(
            item_id__in=Item.objects.values_list("sku", flat=True)
        )
        count_orphans = orphaned_shopitems.count()
        orphaned_shopitems.delete()
        if count_orphans:
            logger.warning(
                "Deleted %d orphaned ShopItem rows with invalid item_id (not matching Item.sku)",
                count_orphans,
            )

    def handle_excel_upload(self):
        """
        Process the uploaded Excel workbook(s) and return the response
        """
        logger.info(
            "handle_excel_upload called for user: %s",
            getattr(self.user, "username", "unknown"),
        )
        # Sanity check: remove orphaned ShopItem rows before processing
        self.cleanup_orphaned_shopitems()
        file_obj = self.request.FILES.get("file")
        if not file_obj or not file_obj.name.endswith(".xlsx"):
            return Response(
                {"detail": "Invalid file format. Please upload an .xlsx file."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        item_field_mapping = {
            "SKU": "sku",
            "Description": "description",
            "Retail Price": "retail_price",
            "Quantity": "quantity",
        }
        shop_item_field_mapping = {
            "Shop User": "shop_user__username",
            "SKU": "item__sku",
            "Description": "item__description",
            "Retail Price": "item__retail_price",
            "Quantity": "quantity",
        }
        excel_item_skus = set()
        unique_shop_items_in_excel = set()
        unique_shop_users_in_excel = set()
        skipped_skus = []
        try:
            logger.info("Starting handle_excel_upload for user: %s", self.user.username)
            workbook = load_workbook(file_obj)
            logger.info("Workbook loaded successfully.")
            with transaction.atomic():
                # Only process sheets that exist; do not error if one is missing
                # Convert custom input format only for the missing sheet, not both
                if "Warehouse Stock" not in workbook.sheetnames:
                    try:
                        converted = self.convert_custom_incoming_format(workbook)
                        if "Warehouse Stock" in converted.sheetnames:
                            workbook = converted
                    except Exception as e:
                        logger.error(
                            "Custom conversion failed for Warehouse Stock: %s",
                            e,
                            exc_info=True,
                        )
                        return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
                if "Warehouse Stock" in workbook.sheetnames:
                    item_sheet = workbook["Warehouse Stock"]
                    headers = [
                        cell.value for cell in next(item_sheet.iter_rows(max_row=1))
                    ]
                    if not all(
                        col in headers
                        for col in ["SKU", "Description", "Retail Price", "Quantity"]
                    ):
                        logger.warning(
                            "Default headers could not be mapped. Consulting custom mappings..."
                        )
                        item_sheet = self.convert_custom_incoming_format(workbook)[
                            "Warehouse Stock"
                        ]
                    for row in item_sheet.iter_rows(min_row=2, values_only=True):
                        data = {
                            item_field_mapping[headers[i]]: value
                            for i, value in enumerate(row)
                            if headers[i] in item_field_mapping
                        }
                        if "retail_price" in data:
                            data["retail_price"] = sanitize_price(data["retail_price"])
                        sku = data.get("sku")
                        if not sku:
                            continue
                        excel_item_skus.add(sku)
                        obj, created = Item.objects.get_or_create(
                            sku=sku, defaults=data
                        )
                        if not created:
                            updated = False
                            for key, value in data.items():
                                if self.field_changed(obj, key, value):
                                    setattr(obj, key, value)
                                    updated = True
                            if obj.is_active is False:
                                obj.is_active = True
                                updated = True
                            if updated:
                                obj.save()
                # --- Deactivate warehouse items not present in the spreadsheet if deletions allowed ---
                if Admin.is_allow_upload_deletions():
                    Item.objects.filter(is_active=True).exclude(sku__in=excel_item_skus).update(is_active=False)
                if "Shop Stock" not in workbook.sheetnames:
                    try:
                        converted = self.convert_custom_incoming_format(workbook)
                        if "Shop Stock" in converted.sheetnames:
                            workbook = converted
                    except Exception as e:
                        logger.error(
                            "Custom conversion failed for Shop Stock: %s",
                            e,
                            exc_info=True,
                        )
                        return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
                if "Shop Stock" in workbook.sheetnames:
                    shop_item_sheet = workbook["Shop Stock"]
                    headers = [
                        cell.value
                        for cell in next(shop_item_sheet.iter_rows(max_row=1))
                    ]
                    if not all(
                        col in headers
                        for col in [
                            "SKU",
                            "Description",
                            "Retail Price",
                            "Quantity",
                            "Shop User",
                        ]
                    ):
                        logger.warning(
                            "Default headers could not be mapped. Consulting custom mappings..."
                        )
                        shop_item_sheet = self.convert_custom_incoming_format(workbook)[
                            "Shop Stock"
                        ]
                        headers = [
                            cell.value
                            for cell in next(shop_item_sheet.iter_rows(max_row=1))
                        ]
                    for row in shop_item_sheet.iter_rows(min_row=2, values_only=True):
                        raw_data = {
                            shop_item_field_mapping[headers[i]]: value
                            for i, value in enumerate(row)
                            if headers[i] in shop_item_field_mapping
                        }
                        shop_username = raw_data.pop("shop_user__username", None)
                        item_sku = raw_data.pop("item__sku", None)
                        if not shop_username or not item_sku:
                            continue
                        unique_shop_items_in_excel.add(item_sku)
                        try:
                            shop_user = User.objects.get(username=shop_username)
                        except User.DoesNotExist:
                            logger.warning(
                                f"Shop user '{shop_username}' not found. Skipping row."
                            )
                            continue
                        unique_shop_users_in_excel.add(shop_user)
                        # --- CHANGED: Create Item if missing, with is_active=False and valid defaults for required fields ---
                        try:
                            item = Item.objects.get(sku=item_sku)
                        except Item.DoesNotExist:
                            # Provide defaults for required fields
                            orig_rp = raw_data.get("item__retail_price", "0.00")
                            try:
                                rp = sanitize_price(orig_rp)
                            except Exception:
                                logger.warning(
                                    "sanitize_price failed for SKU %s value=%r; falling back to 0.00",
                                    item_sku,
                                    orig_rp,
                                )
                                rp = Decimal("0.00")

                            item_defaults = {
                                "description": raw_data.get("item__description", ""),
                                "retail_price": rp,
                                "quantity": raw_data.get("quantity", 0) or 0,
                                "is_active": False,
                            }

                            # Ensure retail_price is a Decimal with 2dp to satisfy model save()
                            try:
                                if not isinstance(item_defaults["retail_price"], Decimal):
                                    item_defaults["retail_price"] = Decimal(item_defaults["retail_price"])
                                item_defaults["retail_price"] = item_defaults["retail_price"].quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                            except Exception as ex:
                                logger.error(
                                    "Failed to coerce retail_price for SKU %s (value=%r): %s",
                                    item_sku,
                                    item_defaults.get("retail_price"),
                                    ex,
                                    exc_info=True,
                                )
                                # As a last resort, set to 0.00
                                item_defaults["retail_price"] = Decimal("0.00")

                            logger.debug("Creating Item with defaults for SKU %s: %r", item_sku, item_defaults)
                            try:
                                item = Item.objects.create(sku=item_sku, **item_defaults)
                            except Exception:
                                logger.error(
                                    "Failed creating Item SKU=%s; defaults=%r",
                                    item_sku,
                                    item_defaults,
                                    exc_info=True,
                                )
                                raise
                            logger.warning(
                                f"Item with SKU '{item_sku}' not found. Created with is_active=False and defaults."
                            )
                        obj, created = ShopItem.objects.get_or_create(
                            shop_user=shop_user, item=item
                        )
                        item_updated = False
                        shop_item_updated = False
                        for key, value in raw_data.items():
                            # Normalize nested item fields (e.g., item__retail_price)
                            if key == "item__retail_price":
                                try:
                                    value = sanitize_price(value)
                                except Exception as exc:
                                    logger.warning(
                                        "Skipping invalid retail_price for SKU %s: %r (%s)",
                                        item_sku,
                                        value,
                                        exc,
                                    )
                                    # record skipped SKU so frontend can be informed
                                    try:
                                        if item_sku not in skipped_skus:
                                            skipped_skus.append(item_sku)
                                    except Exception:
                                        pass
                                    # Don't assign an invalid retail_price to the model; skip this field
                                    continue
                            if key.startswith("item__"):
                                field = key.split("__", 1)[1]
                                if self.field_changed(item, field, value):
                                    setattr(item, field, value)
                                    item_updated = True
                            else:
                                if self.field_changed(obj, key, value):
                                    setattr(obj, key, value)
                                    shop_item_updated = True
                        if item_updated:
                            item.save()
                        if shop_item_updated:
                            obj.save()
                    # --- Delete ShopItems for missing (shop_user, item) only if deletions allowed ---
                    if Admin.is_allow_upload_deletions():
                        excel_shopitem_keys = set()
                        for row in shop_item_sheet.iter_rows(
                            min_row=2, values_only=True
                        ):
                            row_dict = {
                                headers[i]: value
                                for i, value in enumerate(row)
                                if headers[i] in shop_item_field_mapping
                            }
                            shop_username = row_dict.get("Shop User")
                            item_sku = row_dict.get("SKU")
                            if shop_username and item_sku:
                                excel_shopitem_keys.add((shop_username, item_sku))
                        for shop_item in ShopItem.objects.select_related(
                            "shop_user", "item"
                        ).filter(item__isnull=False):
                            key = (shop_item.shop_user.username, shop_item.item.sku)
                            if key not in excel_shopitem_keys:
                                try:
                                    shop_item.delete()
                                except Exception as ex:
                                    logger.error(
                                        f"Failed to delete ShopItem for user '{shop_item.shop_user.username}' and item '{shop_item.item.sku}': {ex}",
                                        exc_info=True,
                                    )
                                    raise
            logger.info("handle_excel_upload completed successfully.")
        except Exception as e:
            logger.error("Error while importing Excel file: %s", str(e), exc_info=True)
            return Response({"detail": "Failed to upload stock data."}, status=400)
        resp_body = {"detail": "Data has been processed according to configuration."}
        if skipped_skus:
            resp_body["skipped_skus"] = skipped_skus
            resp_body["detail"] = "Data processed with some skipped retail_price values. See skipped_skus for list."
        return Response(resp_body, status=200)
